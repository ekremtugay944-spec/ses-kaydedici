"""
Müşteri Excel import/export modülü
"""
import logging
from pathlib import Path
from typing import Tuple
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import db

logger = logging.getLogger(__name__)


COLUMNS = [
    ("name", "İsim (zorunlu)", 30),
    ("phone", "Telefon", 18),
    ("company", "Firma", 30),
    ("city", "Şehir", 15),
    ("address", "Adres", 40),
    ("tax_id", "Vergi No", 15),
    ("segment", "Segment", 15),
    ("notes", "Notlar", 40),
]


def create_template(path: Path):
    """Boş import şablonu oluştur"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Müşteriler"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    for i, (key, label, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=i, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = width

    # Örnek satır
    examples = [
        "Ahmet Bey - Bursa Spor",
        "+90 532 123 45 67",
        "Bursa Spor Mağazası",
        "Bursa",
        "Atatürk Cd. No:12",
        "1234567890",
        "Toptan",
        "Düzenli ödeyici",
    ]
    for i, val in enumerate(examples, start=1):
        ws.cell(row=2, column=i, value=val)

    # Talimat sayfası
    ws2 = wb.create_sheet("Talimatlar")
    instructions = [
        "MÜŞTERİ IMPORT ŞABLONU",
        "",
        "1. 'Müşteriler' sekmesinde 2. satırdaki ÖRNEĞİ SİL.",
        "2. Müşterilerini ekle (her satır bir müşteri).",
        "3. SADECE 'İsim' zorunlu, diğerleri boş kalabilir.",
        "4. Telefon formatı: +90 5XX XXX XX XX (WhatsApp linki için).",
        "5. Dosyayı kaydet → bota /import_musteri komutu ile dosyayı yolla.",
        "",
        "İSİM YAZIM KURALI:",
        "- 'Ahmet' yerine 'Ahmet Bey - Bursa Spor' tercih et (daha sonra aramada kolay).",
        "- Bu sayede 'Ahmet'i aradım' dediğinde bot doğru müşteriyi bulur.",
        "",
        "SEGMENT ÖRNEKLERİ:",
        "- Toptan, Perakende, Online, VIP, Düşük öncelik",
    ]
    for i, line in enumerate(instructions, start=1):
        ws2.cell(row=i, column=1, value=line)
    ws2.column_dimensions["A"].width = 80

    wb.save(path)


def import_from_excel(path: Path) -> Tuple[int, int, list]:
    """
    Excel'den müşterileri içe aktar.
    Dönüş: (eklenen_sayı, güncellenen_sayı, hata_listesi)
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    added = 0
    updated = 0
    errors = []

    # Başlıkları bul (1. satır)
    headers = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            label = str(cell.value).strip()
            # Etiketten anahtara çevir
            for key, expected_label, _ in COLUMNS:
                if expected_label.lower().split()[0] in label.lower():
                    headers[col_idx] = key
                    break

    # Veri satırlarını oku
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue

        record = {}
        for col_idx, value in enumerate(row, start=1):
            key = headers.get(col_idx)
            if key and value is not None:
                record[key] = str(value).strip()

        name = record.pop("name", None)
        if not name:
            errors.append(f"Satır {row_idx}: İsim boş, atlandı")
            continue

        try:
            existed_before = db.get_customer(name) is not None
            db.upsert_customer(name, **record)
            if existed_before:
                updated += 1
            else:
                added += 1
        except Exception as e:
            errors.append(f"Satır {row_idx} ({name}): {e}")

    wb.close()
    return added, updated, errors


def export_to_excel(path: Path):
    """Mevcut müşterileri Excel'e aktar"""
    customers = db.list_customers(limit=10000)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Müşteriler"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    for i, (key, label, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=i, value=label)
        cell.fill = header_fill
        cell.font = header_font
        ws.column_dimensions[cell.column_letter].width = width

    for row_idx, c in enumerate(customers, start=2):
        for col_idx, (key, _, _) in enumerate(COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=c.get(key, ""))

    wb.save(path)
    return len(customers)
